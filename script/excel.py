import xlrd
from script.styling import *

from openpyxl import Workbook

from script.repositories.place_list import Places
from script.repositories.people_list import People
from script.models.person import Person
from script.models.place import Place
from script.models.day import Day


def convert(file):

    #region READING_FUNC
    def get_rows_cols():
        """
        Count real value of columns and rows,
        How I calculate this? I simply check if cell have border,
        :return: cols number
        """
        final_row = 0
        final_col = 0

        for col in range(sheet.ncols):
            cell_xf = book.xf_list[sheet.cell_xf_index(5, col)]
            if cell_xf.border.right_line_style == 0:
                final_col = col - 1
                break

        for row in range(5, sheet.nrows):
            cell_xf = book.xf_list[sheet.cell_xf_index(row, 0)]
            if cell_xf.border.bottom_line_style == 0:
                final_row = row - 1
                break

        return final_row, final_col

    def get_val_color(row, col):
        cell = sheet.cell(row, col).value
        color = book.xf_list[sheet.cell_xf_index(row, col)].background.background_colour_index
        return cell, color

    def fill_places():
        list_of_places = Places()
        for i in range(1, cols):
            if sheet.cell(0, i).value != "":
                list_of_places.add_place(Place(sheet.cell(0, i).value, book.xf_list[
                    sheet.cell_xf_index(0, i)].background.background_colour_index))
        return list_of_places

    def fill_people():
        people_list = People()

        for person in range(4, rows, 2):
            count_week = 1
            temp_person = Person(sheet.cell(person, 0).value)
            for day in range(1, len(days) + 1):
                if days[day - 1] == "":
                    count_week += 1
                else:
                    day_cell, day_color = get_val_color(person, day)
                    info_day = Day(day_cell, sheet.cell(person + 1, day).value, places.get_place(day_color), count_week)
                    temp_person.add_day(info_day)
            people_list.add_person(temp_person)
        return people_list

    def get_days():
        list_of_days = [sheet.cell(2, col).value for col in range(1, cols)]
        while list_of_days[-1] is "":
            list_of_days.pop()
        return list_of_days
    #endregion
    #region WRITING_FUNC
    def weekend_colored(worksheet, col, row):
        for (week) in (worksheet.iter_cols(min_row=row, max_row=row, min_col=col + 1,
                                           max_col=col + peoples.count() * 2 + 2)):
            for cell_week in week:
                cell_week.fill = weekend_fill

    def write_days(worksheet, col=0):
        if col != 0:
            col *= peoples.count() * 2 + 3
        cell1, cell2 = worksheet.cell(1, col + 1), worksheet.cell(1, col + 2)
        worksheet.column_dimensions[cell1.column].width = 2.5
        worksheet.column_dimensions[cell2.column].width = 2.5

        day_no = 1
        for (cell, day) in zip(worksheet.iter_rows(min_row=3, max_row=len(days) + 2, min_col=col + 1, max_col=col + 2),
                               days):
            cell[0].value, cell[1].value = day_no, day
            cell[0].border = cell[1].border = thin_border
            if day != "":
                day_no += 1
                if day in ["So", "N"]:
                    weekend_colored(worksheet, col, cell[0].row)
            else:
                sum_rows.add(cell[0].row)
                cell[0].value = "Suma"

    def write_header(worksheet, s_col=0):
        if s_col != 0:
            s_col *= peoples.count() * 2 + 3
        i = 0
        for col in (
        worksheet.iter_cols(min_row=2, max_row=2, min_col=s_col + 3, max_col=s_col + peoples.count() * 2 + 2)):
            cell = col[0]
            cell.fill, cell.font, cell.alignment = header_fill, header_font, center
            if i % 2 == 0:
                cell.value = peoples.get_person_name_by_id(i // 2)
                cell.border = upper_left_border
            else:
                cell.value = 'H'
                cell.border = upper_right_border
            i += 1

    def write_table(worksheet):
        i = 0
        for (work) in (
        worksheet.iter_cols(min_row=3, max_row=len(days) + 2, min_col=3, max_col=peoples.count() * 2 + 2)):
            cell = work[0]
            person = peoples.get_person_by_id(i)
            if cell.col_idx % 2 != 0:
                day_no = 0
                for (day) in work:
                    day.border = left_border
                    day.alignment = center
                    if day.row in sum_rows:
                        day_no -= 1
                    elif person.get_place_of_day(day_no) == 'biuro':
                        day.value = person.get_time_of_day(day_no)
                    elif person.get_time_of_day(day_no) == 'wolne':
                        day.value = person.get_time_of_day(day_no)
                        day.fill = wolne_fill
                    elif person.get_place_of_day(day_no) in places.get_places_names():
                        day.value = person.get_place_of_day(day_no).upper()
                        day.fill = wyspa_fill
                    else:
                        day.value = person.get_time_of_day(day_no)

                    day_no += 1
            else:  # HOURS
                worksheet.column_dimensions[work[0].column].width = 2.5
                i += 1
                day_no = 0
                for (day) in work:
                    day.border = right_border
                    day.alignment = center
                    if day.row in sum_rows:
                        day_no -= 1
                    else:
                        day.value = person.get_hours_of_day(day_no)
                        if person.get_time_of_day(day_no) == 'wolne':
                            day.fill = wolne_fill
                        elif person.get_place_of_day(day_no) in places.get_places_names():
                            day.fill = wyspa_fill

                    day_no += 1

    def write_table_wyspa(worksheet, wyspa, s_col=1):

        s_col *= peoples.count() * 2 + 3
        worksheet.cell(1, s_col + 3, wyspa)
        i = y = 0
        for (work) in (worksheet.iter_cols(min_row=3, max_row=len(days) + 2, min_col=s_col + 3,
                                           max_col=s_col + peoples.count() * 2 + 2)):
            person = peoples.get_person_by_id(i)
            if y % 2 == 0:
                day_no = 0
                for (day) in work:
                    day.border = left_border
                    day.alignment = center
                    if day.row in sum_rows:
                        day_no -= 1
                    elif person.get_place_of_day(day_no) == wyspa:
                        day.value = person.get_time_of_day(day_no)
                    day_no += 1
                    y += 1
            else:  # HOURS
                i += 1
                day_no = 0
                for (day) in work:
                    day.border = right_border
                    day.alignment = center
                    if day.row in sum_rows:
                        day_no -= 1
                    elif person.get_place_of_day(day_no) == wyspa:
                        day.value = person.get_hours_of_day(day_no)
                    day_no += 1
                y += 1
    #endregion

    #region READ_FILE
    book = xlrd.open_workbook(file, formatting_info=True)
    sheet = book.sheet_by_index(0)

    rows, cols = get_rows_cols()
    places = fill_places()
    days = get_days()
    peoples = fill_people()
    #endregion
    #region WRITE_FILE
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Salon Łęczycka godziny pracy: pon-pt. 9.00-18.00"
    ws['A2'] = "Mnth-Year"
    sum_rows = set()

    # ...........CONFIGURING...............
    for x in range(1, 60):
        for y in range(1, 100):
            cell = ws.cell(row=x, column=y)
            cell.fill, cell.font = white_fill, def_font
    rd = ws.row_dimensions[2]
    rd.height = 30

    write_days(ws)
    write_header(ws)
    write_table(ws)

    write_days(ws, 1)
    write_header(ws, 1)
    write_table_wyspa(ws, 'focus')

    write_days(ws, 2)
    write_header(ws, 2)
    write_table_wyspa(ws, 'arkady', 2)

    return wb
    #endregion



from openpyxl.styles import Border, Side, Alignment
from openpyxl.styles import Color, Border, Font, Side, PatternFill

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

upper_left_border = Border(left=Side(style='medium'),
                           right=Side(style='thin'),
                           top=Side(style='medium'),
                           bottom=Side(style='thin'))

left_border = Border(left=Side(style='medium'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

right_border = Border(left=Side(style='thin'),
                           right=Side(style='medium'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

upper_right_border = Border(left=Side(style='thin'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='thin'))

def_font = Font(size=8)
header_font = Font(size=8, bold=True, color='ffffff')

center = Alignment(horizontal='center', vertical='center', wrap_text=True)



white_fill = PatternFill(fill_type='solid',
                         start_color='ffffff',
                         end_color='ffffff')

header_fill = PatternFill(fill_type='solid',
                          start_color='938953',
                          end_color='938953')

date_fill = PatternFill(fill_type='solid',
                        start_color='92d050',
                        end_color='92d050')

wyspa_fill = PatternFill(fill_type='solid',
                        start_color='9966ff',
                        end_color='9966ff')

wolne_fill = PatternFill(fill_type='solid',
                        start_color='ff0000',
                        end_color='ff0000')

weekend_fill = PatternFill(fill_type='solid',
                        start_color='CCC0DA',
                        end_color='CCC0DA')





